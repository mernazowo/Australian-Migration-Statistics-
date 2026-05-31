#!/usr/bin/env python3
"""Export clean CSV tables from migration_trends_statistical_package_2024_25.xlsx."""

from __future__ import annotations

import csv
import re
from pathlib import Path

import openpyxl

XLSX = Path(
    r"d:\Third Year\FIT2179 Data Visualisation\Assignment 2"
    r"\migration_trends_statistical_package_2024_25.xlsx"
)
OUT = Path(__file__).resolve().parent.parent / "data"

FOOTNOTE_START = re.compile(
    r"^(?:\d+\.|Note:|Source|Click here|Table |\s*$)",
    re.IGNORECASE,
)


def clean_cell(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def clean_header(value):
    text = clean_cell(value)
    text = re.sub(r"\d+$", "", text).strip()
    text = text.replace("People's Republic of China", "China")
    return text


def row_values(ws, row_idx, max_col=None):
    max_col = max_col or ws.max_column
    return [ws.cell(row_idx, col).value for col in range(1, max_col + 1)]


def write_csv(path: Path, headers, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)


def find_data_end(ws, start_row, year_col=2):
    row = start_row
    while row <= ws.max_row:
        year = ws.cell(row, year_col).value
        if year is None or str(year).strip() == "":
            break
        label = clean_cell(year)
        if FOOTNOTE_START.match(label):
            break
        row += 1
    return row - 1


def export_sheet_1_0(ws):
    headers = [clean_header(v) for v in row_values(ws, 2)[1:7]]
    rows = []
    end = find_data_end(ws, 3)
    for r in range(3, end + 1):
        vals = row_values(ws, r)[1:7]
        rows.append([clean_cell(v) for v in vals])
    write_csv(OUT / "migration_program.csv", headers, rows)


def export_sheet_1_1(ws):
    headers = [clean_header(v) for v in row_values(ws, 3)[1:20]]
    rows = []
    end = find_data_end(ws, 4)
    for r in range(4, end + 1):
        vals = row_values(ws, r)[1:20]
        rows.append([clean_cell(v) for v in vals])
    write_csv(OUT / "streams.csv", headers, rows)


def export_country_matrix(ws, out_name):
    headers = [clean_header(v) for v in row_values(ws, 2)[1:]]
    while headers and headers[-1] == "":
        headers.pop()
    rows = []
    end = find_data_end(ws, 3)
    for r in range(3, end + 1):
        vals = row_values(ws, r)[1 : len(headers) + 1]
        rows.append([clean_cell(v) for v in vals])
    write_csv(OUT / out_name, headers, rows)


def export_sheet_1_11(ws):
    raw_headers = row_values(ws, 2)[1:]
    headers = [clean_header(v) for v in raw_headers if clean_header(v)]
    if headers and headers[-1] == "Total":
        headers = headers[:-1]
    rows = []
    end = find_data_end(ws, 3, year_col=2)
    for r in range(3, end + 1):
        vals = row_values(ws, r)[1 : len(headers) + 1]
        rows.append([clean_cell(v) for v in vals])
    write_csv(OUT / "occupations.csv", headers, rows)


def export_sheet_3_0(ws):
    headers = [
        "Year",
        "Refugee",
        "Special Humanitarian Program",
        "Offshore resettlement total",
        "Offshore % of total Program",
        "Onshore protection total",
        "Onshore % of total Program",
        "Total Humanitarian Program grants",
    ]
    rows = []
    end = find_data_end(ws, 4)
    for r in range(4, end + 1):
        vals = row_values(ws, r)[1:9]
        rows.append([clean_cell(v) for v in vals])
    write_csv(OUT / "humanitarian.csv", headers, rows)


def export_nom_sheet(ws, out_name):
    headers = [clean_header(v) for v in row_values(ws, 3)[1:18]]
    rows = []
    end = find_data_end(ws, 4)
    for r in range(4, end + 1):
        vals = row_values(ws, r)[1:18]
        rows.append([clean_cell(v) for v in vals])
    write_csv(OUT / out_name, headers, rows)


def export_sheet_6_0(ws):
    headers = [clean_header(v) for v in row_values(ws, 2)[1:12]]
    rows = []
    for r in range(3, 5):
        vals = row_values(ws, r)[1:12]
        if clean_cell(vals[0]):
            rows.append([clean_cell(v) for v in vals])
    write_csv(OUT / "citizenship.csv", headers, rows)


def export_sheet_7_0(ws):
    headers = ["Metric", "Australian-born", "Established migrants", "Recent migrants", "Total"]
    rows = []
    for r in range(3, 8):
        label = clean_cell(ws.cell(r, 2).value)
        if not label or FOOTNOTE_START.match(label):
            break
        vals = row_values(ws, r)[2:6]
        rows.append([label, *[clean_cell(v) for v in vals]])
    write_csv(OUT / "employment_growth.csv", headers, rows)


def export_sheet_7_1(ws):
    headers = [
        "Years in Australia",
        "Skill stream participation",
        "Family stream participation",
        "Australian-born participation",
        "Skill stream unemployment",
        "Family stream unemployment",
        "Australian-born unemployment",
    ]
    rows = []
    end = find_data_end(ws, 4)
    for r in range(4, end + 1):
        vals = row_values(ws, r)[1:8]
        rows.append([clean_cell(v) for v in vals])
    write_csv(OUT / "unemployment_integration.csv", headers, rows)


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=False, data_only=True)
    export_sheet_1_0(wb["1.0"])
    export_sheet_1_1(wb["1.1"])
    export_country_matrix(wb["1.3"], "employer_sponsored_countries.csv")
    export_country_matrix(wb["1.8"], "state_nominated_countries.csv")
    export_sheet_1_11(wb["1.11"])
    export_sheet_3_0(wb["3.0"])
    export_nom_sheet(wb["5.0"], "nom_arrivals.csv")
    export_nom_sheet(wb["5.1"], "nom_departures.csv")
    export_sheet_6_0(wb["6.0"])
    export_sheet_7_0(wb["7.0"])
    export_sheet_7_1(wb["7.1"])
    print("Exported CSV files to", OUT)
    for p in sorted(OUT.glob("*.csv")):
        print(" ", p.name)


if __name__ == "__main__":
    main()
