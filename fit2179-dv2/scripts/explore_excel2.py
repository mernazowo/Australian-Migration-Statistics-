import openpyxl
from pathlib import Path

XLSX = Path(r"d:\Third Year\FIT2179 Data Visualisation\Assignment 2\migration_trends_statistical_package_2024_25.xlsx")

for name in ["1.1", "1.11", "1.0", "3.0"]:
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb[name]
    print(f"\n{'='*60}\nSHEET {name}  rows={ws.max_row}")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), start=1):
        vals = ["" if v is None else v for v in row]
        if any(str(v).strip() for v in vals):
            print(f"{i:3}: {vals[:15]}")
