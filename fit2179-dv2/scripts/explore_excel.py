import openpyxl
from pathlib import Path

XLSX = Path(r"d:\Third Year\FIT2179 Data Visualisation\Assignment 2\migration_trends_statistical_package_2024_25.xlsx")
SHEETS = ["1.0", "1.1", "1.3", "1.8", "1.11", "3.0", "5.0", "5.1", "6.0", "7.0", "7.1"]

wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
for name in SHEETS:
    ws = wb[name]
    print(f"\n{'='*60}\nSHEET {name}  rows={ws.max_row} cols={ws.max_column}")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=35, values_only=True), start=1):
        vals = ["" if v is None else v for v in row]
        if any(str(v).strip() for v in vals):
            print(f"{i:3}: {vals[:12]}")
